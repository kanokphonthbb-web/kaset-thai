#!/usr/bin/env python3

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path('/Users/bob/Downloads/kasetkonthai_10000_seo_aeo_ai_search_titles (1).xlsx')
OUTPUT_DIR = ROOT / 'data' / 'affiliate-v2'
OUTPUT = OUTPUT_DIR / 'source-workbook.json'


def serializable(value):
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return value


workbook = load_workbook(SOURCE, read_only=True, data_only=False)
sheets = []
for worksheet in workbook.worksheets:
    rows = [
        [serializable(value) for value in row]
        for row in worksheet.iter_rows(values_only=True)
    ]
    while rows and all(value is None for value in rows[-1]):
        rows.pop()
    max_columns = max((len(row) for row in rows), default=0)
    sheets.append({
        'name': worksheet.title,
        'rows': len(rows),
        'columns': max_columns,
        'values': rows,
    })

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps({
    'source_path': str(SOURCE),
    'extraction_mode': 'openpyxl_read_only_data_only_false',
    'sheets': sheets,
}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

print(json.dumps({
    'output': str(OUTPUT),
    'sheets': [
        {
            'name': sheet['name'],
            'rows': sheet['rows'],
            'columns': sheet['columns'],
            'preview': sheet['values'][:4],
        }
        for sheet in sheets
    ],
}, ensure_ascii=False, indent=2))
